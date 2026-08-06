# -*- coding: utf-8 -*-
"""BRFplus Configuration Workbook: decision table content ready for BRFplus Excel import."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HDR_FILL = PatternFill("solid", fgColor="1F4E79")
ALT_FILL = PatternFill("solid", fgColor="DEEAF6")
NOTE_FILL = PatternFill("solid", fgColor="FFF2CC")
HDR_FONT = Font(name="Arial", size=10, bold=True, color="FFFFFF")
BODY = Font(name="Arial", size=10)
BOLD = Font(name="Arial", size=10, bold=True)
TITLE = Font(name="Arial", size=13, bold=True, color="1F4E79")
thin = Side(style="thin", color="BBBBBB")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = openpyxl.Workbook()

def sheet(name, title, note, headers, rows, widths):
    ws = wb.create_sheet(name)
    ws["A1"] = title; ws["A1"].font = TITLE
    ws["A2"] = note; ws["A2"].font = Font(name="Arial", size=9, italic=True, color="595959")
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max(2, len(headers)))
    ws.row_dimensions[2].height = 28
    r0 = 4
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=r0, column=c, value=h)
        cell.font = HDR_FONT; cell.fill = HDR_FILL; cell.border = BORDER
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    for ri, row in enumerate(rows):
        for c, v in enumerate(row, 1):
            cell = ws.cell(row=r0 + 1 + ri, column=c, value=v)
            cell.font = BODY; cell.border = BORDER
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if ri % 2 == 1: cell.fill = ALT_FILL
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = ws.cell(row=r0 + 1, column=1)
    return ws

# ---------- Summary ----------
ws = wb.active; ws.title = "00_Summary"
ws["A1"] = "ZEC_PROGRAM_ELIGIBILITY — BRFplus Object Inventory & Configuration Workbook"; ws["A1"].font = TITLE
ws["A3"] = ("Companion to 'Enrollment Center – BRFplus Detailed Configuration'. Decision-table sheets contain the "
            "complete productive content, structured for BRFplus Workbench import (open the decision table → "
            "Additional Actions → Import From Excel → map columns 1:1; header row = condition/result columns). "
            "All values in this workbook are configuration content defined in the design documents — no formulas.")
ws["A3"].font = Font(name="Arial", size=9, italic=True); ws["A3"].alignment = Alignment(wrap_text=True, vertical="top")
ws.merge_cells("A3:F3"); ws.row_dimensions[3].height = 44
inv = [
 ("Object Type","Count","Objects","Sheet"),
 ("Application","1","ZEC_PROGRAM_ELIGIBILITY","—"),
 ("Functions","2","F_EC_ELIGIBILITY_ALL (event mode); F_EC_DEENROLL_CHECK (functional mode)","10_Functions"),
 ("Rulesets","28","One per program, assigned to F_EC_ELIGIBILITY_ALL","11_Rulesets"),
 ("Decision tables","7","DT_RATECLASS_MATRIX; DT_PROGRAM_CRITERIA; DT_PROGRAM_EXCLUSIONS; DT_WAITING_PERIODS; DT_PROGRAM_WARNINGS; DT_PREACTIVE_INFO; DT_EXIT_RULES","01–07"),
 ("Reusable expressions","7","EX_RATE_CLASS, EX_WAITING_PERIOD, EX_PROGRAM_EXCLUSION, EX_AMI_CHECK, EX_HIST_MONTHS_OK, EX_SEVERITY_MODE, EX_MSG_APPEND","—"),
 ("Data objects (structures/tables)","4","CTX_EC_ELIG, CTX_EC_CA, CTX_EC_HIST, RES_EC_ELIG","—"),
 ("Message class","1 (33 messages)","ZEC_ELIG","12_Messages"),
]
for ri, row in enumerate(inv):
    for ci, v in enumerate(row, 1):
        cell = ws.cell(row=5 + ri, column=ci, value=v)
        cell.border = BORDER
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        if ri == 0: cell.font = HDR_FONT; cell.fill = HDR_FILL
        else:
            cell.font = BODY
            if ri % 2 == 0: cell.fill = ALT_FILL
ws["A15"] = ("Design decision: per-program condition content is consolidated into criteria-matrix tables "
             "(DT_PROGRAM_CRITERIA / DT_PROGRAM_WARNINGS) — one ROW per program instead of one TABLE per program. "
             "The naive per-program alternative would require 40 decision tables (4 shared + 20 hard-stop + 16 warning); "
             "the consolidated model needs 7 and keeps maintenance in two business-owned matrices.")
ws["A15"].font = Font(name="Arial", size=9); ws["A15"].fill = NOTE_FILL
ws["A15"].alignment = Alignment(wrap_text=True, vertical="top")
ws.merge_cells("A15:F16")
for i, w in enumerate([26, 12, 60, 14, 10, 10], 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# ---------- 01 DT_RATECLASS_MATRIX ----------
sheet("01_DT_RATECLASS_MATRIX", "DT_RATECLASS_MATRIX",
 "Condition: RATE_CATEGORY (pattern match). Result: RATE_CLASS. Used by EX_RATE_CLASS. Maintain patterns per the client's rate catalog at realization.",
 ["RATE_CATEGORY (pattern)","RATE_CLASS (result)"],
 [["RS*","RES"],["RES*","RES"],["E-1*","RES"],["G-1*","RES"],
  ["GS*","COM"],["COM*","COM"],["E-2*","COM"],["LGS*","COM"],["C&I*","COM"],
  ["* (fallback)","RES"]],
 [30, 20])

# ---------- 02 DT_PROGRAM_CRITERIA ----------
crit_rows = [
 ["DPP","Dynamic Peak Pricing","RES","X","","","","","","", "","","","","Y"],
 ["PTR","Peak-Time Rebate","RES","X","X","","","","","","","","","","Y"],
 ["CPP","Critical Peak Pricing","RES","X","","","","","","","","","","","Y"],
 ["TOU","Time-of-Use Rate","","X","","","","","","","","","","","Y"],
 ["PPS","Peak Power Savers (AC Cycling)","RES","","","","","","","","","","","","I"],
 ["THERM","Smart Thermostat","RES","","","","","","","","","","","","Y"],
 ["EVMC","EV Managed Charging","","","","","","","","","","","","","Y"],
 ["EEREB","Energy-Efficiency Rebates","","","","","","","","","","","","","Y"],
 ["CSDD","Customer Selects Due Date","","","","X","","","","","","","","","Y"],
 ["EBILL","eBill / Paperless","","","","","","","","","","","","","Y"],
 ["DPA","Deferred Payment Arrangement","","","","","","","X","X","","","","","I"],
 ["BB","Budget Billing","","","","X","","","","","12","","","","I"],
 ["PREPAY","PrePay","RES","X","","","X","X","X","","","","","","Y"],
 ["AUTOPAY","AutoPay","","","","","","","","","","","","","Y"],
 ["PEXT","Payment Extension","","","","","","","","","","","","X","I"],
 ["SUMBILL","Summary/Consolidated Billing","","","","","","","","","","","2","","N"],
 ["TPN","Third-Party Notification","","","","","","","","","","","","","Y"],
 ["LIHEAP","LIHEAP Assistance","RES","","","","","","","","","","","","Y"],
 ["PIPP","Percentage of Income Plan","RES","","","","","","","","","","","","Y"],
 ["AMP","Arrearage Management","","","","","","","","X","","","","","I"],
 ["CARE","CARE / FERA Discount","RES","","","","","","","","","","","","Y"],
 ["MED","Medical Alert","","","","","","","","","","","","","Y"],
 ["GREEN","Green Power","","","","","","","","","","","","","Y"],
 ["CSOL","Community Solar","","","","","","","","","","","","","Y"],
 ["ALERTS","Usage/High-Bill Alerts","","X","X","","","","","","","","","","Y"],
 ["INTR","Interruptible Tariff","COM","","","","","","","","","100","","","N"],
 ["ADR","Automated DR (OpenADR)","COM","","","","","","","","","","","","N"],
 ["CDR","C&I Demand Response","COM","","","","","","","","","50","","","N"],
]
sheet("02_DT_PROGRAM_CRITERIA", "DT_PROGRAM_CRITERIA (criteria matrix — one row per program)",
 "Generic hard-stop rules read this matrix: each violated criterion appends its fixed message — RATE_CLASS→002/003, AMI→044, AMI_COMM→018, PREPAY block→055/056, MEDICAL→052, BB→056, DPA→012/057, ARREARS→071/072, HISTORY→062(I), LOAD→080, CA_COUNT→074, EXT_ACTIVE→073. PREACTIVE: Y=enrollable at move-in, I=informational only (DT_PREACTIVE_INFO), N=hidden in move-in mode.",
 ["PROGRAM_ID","PROGRAM_NAME","REQ_RATE_CLASS","REQ_AMI","REQ_AMI_COMM","BLOCK_IF_PREPAY","BLOCK_IF_MEDICAL","BLOCK_IF_BB","BLOCK_IF_DPA","REQ_ARREARS","MIN_HISTORY_MONTHS","MIN_LOAD_KW","MIN_CA_COUNT","BLOCK_IF_EXT_ACTIVE","PREACTIVE"],
 crit_rows,
 [10, 26, 12, 8, 11, 12, 13, 9, 10, 11, 14, 11, 12, 14, 10])

# ---------- 03 DT_PROGRAM_EXCLUSIONS ----------
sheet("03_DT_PROGRAM_EXCLUSIONS", "DT_PROGRAM_EXCLUSIONS (mutual exclusions)",
 "Condition: requested program + active participation (T_ACTIVE). Result: severity + ZEC_ELIG message.",
 ["REQUESTED_PROGRAM","BLOCKING_ACTIVE_PROGRAM","SEVERITY","MSGNO"],
 [["PREPAY","MED","E","052"],["PREPAY","BB","E","056"],["PREPAY","DPA","E","057"],
  ["BB","PREPAY","E","056"],["CSDD","PREPAY","E","055"],["DPA","DPA","E","012"],
  ["DPP","CPP","E","058"],["CPP","DPP","E","058"],
  ["TOU","DPP","W","340"],["TOU","CPP","W","340"],
  ["GREEN","CSOL","W","341"],["CSOL","GREEN","W","341"]],
 [22, 26, 10, 8])

# ---------- 04 DT_WAITING_PERIODS ----------
sheet("04_DT_WAITING_PERIODS", "DT_WAITING_PERIODS",
 "Condition: program + latest prior participation outcome (C=Completed, D=De-enrolled, X=Cancelled, F=Default). Result: waiting months + message. Evaluated by EX_WAITING_PERIOD against T_HISTORY.",
 ["PROGRAM_ID","PRIOR_OUTCOME","WAITING_MONTHS","MSGNO"],
 [["CPP","D","6","045"],["DPP","D","6","045"],["PTR","D","3","045"],
  ["PPS","D","12","045"],["PREPAY","D","3","045"],
  ["AMP","F","24","046"],["DPA","F","12","046"]],
 [12, 14, 16, 8])

# ---------- 05 DT_PROGRAM_WARNINGS ----------
sheet("05_DT_PROGRAM_WARNINGS", "DT_PROGRAM_WARNINGS (one row per warning condition)",
 "Condition: program + condition code (evaluated from CTX_EC_CA / BP data) + threshold. Result: message (severity W). Warning rows never block; the UI offers supervisor override where policy requires it.",
 ["PROGRAM_ID","CONDITION_CODE","THRESHOLD","MSGNO","CONDITION SOURCE"],
 [["TOU","ACTIVE_DYNRATE","","340","T_ACTIVE contains DPP/CPP"],
  ["THERM","NO_WIFI","","342","Capture form / premise survey flag"],
  ["EVMC","TELEMATICS_UNVERIFIED","","320","Device compatibility list"],
  ["EEREB","INVOICE_VERIFY","","321","Always (process rule)"],
  ["CSDD","RETURNED_PAYMENTS","2","031","CTX_EC_CA.RETURNED_PAYMENTS_12M"],
  ["EBILL","NO_EMAIL","","322","BP communication data"],
  ["DPA","BROKEN_PLANS","2","330","CTX_EC_CA.BROKEN_PLANS_24M"],
  ["AUTOPAY","RETURNED_PAYMENTS","2","031","CTX_EC_CA.RETURNED_PAYMENTS_12M"],
  ["PEXT","EXT_COUNT_12M","2","331","CTX_EC_CA.PAYMENT_EXT_12M"],
  ["LIHEAP","INCOME_VERIFY","","332","Always (provisional enrollment)"],
  ["PIPP","INCOME_VERIFY","","332","Always"],
  ["CARE","INCOME_VERIFY","","332","Always (self-certification)"],
  ["MED","CERT_DUE","30","333","Always (certification deadline days)"],
  ["GREEN","STACKING_CSOL","","341","T_ACTIVE contains CSOL"],
  ["CSOL","CAPACITY_WAITLIST","","334","Array capacity service"],
  ["ADR","TELEMETRY_CERT","","335","Telemetry certification status"]],
 [12, 24, 11, 8, 36])

# ---------- 06 DT_PREACTIVE_INFO ----------
sheet("06_DT_PREACTIVE_INFO", "DT_PREACTIVE_INFO (mode P informational rows)",
 "For programs with PREACTIVE = I in DT_PROGRAM_CRITERIA: informational message shown in the Move-In context (severity I). Programs with PREACTIVE = N are omitted from move-in evaluation entirely.",
 ["PROGRAM_ID","MSGNO","MESSAGE INTENT"],
 [["DPA","090","Payment plans require an active contract with billing history"],
  ["BB","062","Requires 12 months billing history at premise"],
  ["PPS","093","Device installation scheduled after contract activation"],
  ["PEXT","090","Requires an active contract"],
  ["AMP","090","Requires an active contract with billing history"]],
 [12, 8, 52])

# ---------- 07 DT_EXIT_RULES ----------
sheet("07_DT_EXIT_RULES", "DT_EXIT_RULES (de-enrollment)",
 "Used by F_EC_DEENROLL_CHECK. Condition: program + exit condition code. Result: severity + EC_EXIT message. E blocks de-enrollment; W informs.",
 ["PROGRAM_ID","EXIT_CONDITION","SEVERITY","MSGNO (EC_EXIT)","CONDITION SOURCE"],
 [["DPA","OPEN_INSTPLAN_BALANCE","E","007","FI-CA installment plan open amount > 0"],
  ["BB","TRUEUP_PENDING","W","010","Unbilled true-up at annual review"],
  ["THERM","OPEN_SERVICE_ORDER","E","011","Device removal/install order open"],
  ["PPS","OPEN_SERVICE_ORDER","E","011","Switch removal order open"],
  ["TOU","MIN_PARTICIPATION","E","012","Enrolled < 12 months"],
  ["INTR","CONTRACT_TERM","E","013","Within contract term"],
  ["PREPAY","NEGATIVE_BALANCE","E","014","Prepay balance < 0"]],
 [12, 24, 10, 14, 40])

# ---------- 10 Functions ----------
sheet("10_Functions", "Functions",
 "Both functions live in application ZEC_PROGRAM_ELIGIBILITY. GUIDs are stored in ZEC_FDT_IDS after transport — never hardcoded.",
 ["FUNCTION","MODE","CONTEXT","RESULT","PROCESSING"],
 [["F_EC_ELIGIBILITY_ALL","Event Mode","CTX_EC_ELIG (BP, MODE, MOVEIN_ID, PROGRAM_FILTER, T_CA, T_HISTORY, T_ACTIVE, REQUEST_DATE)","RES_EC_ELIG (table)","Executes all 28 assigned rulesets; ruleset preconditions filter to requested programs; rules loop over T_CA"],
  ["F_EC_DEENROLL_CHECK","Functional Mode","PROGRAM_ID, CONTRACT_ACCOUNT, enrollment dates, CTX_EC_CA row","RES_EC_ELIG","Top expression evaluates DT_EXIT_RULES + FI-CA balance lookups"]],
 [22, 12, 42, 16, 42])

# ---------- 11 Rulesets ----------
programs = [
 ("RS_DPP_INT_VALIDATION","DPP","Dynamic Peak Pricing (migrated legacy ruleset name)"),
 ("RS_PTR_VALIDATION","PTR","Peak-Time Rebate"),("RS_CPP_VALIDATION","CPP","Critical Peak Pricing"),
 ("RS_TOU_VALIDATION","TOU","Time-of-Use"),("RS_PPS_VALIDATION","PPS","Peak Power Savers"),
 ("RS_THERM_VALIDATION","THERM","Smart Thermostat"),("RS_EVMC_VALIDATION","EVMC","EV Managed Charging"),
 ("RS_EEREB_VALIDATION","EEREB","EE Rebates"),("RS_CSDD_VALIDATION","CSDD","Customer Selects Due Date"),
 ("RS_EBILL_VALIDATION","EBILL","eBill"),("RS_DPA_VALIDATION","DPA","Deferred Payment Arrangement"),
 ("RS_BB_VALIDATION","BB","Budget Billing"),("RS_PREPAY_VALIDATION","PREPAY","PrePay"),
 ("RS_AUTOPAY_VALIDATION","AUTOPAY","AutoPay"),("RS_PEXT_VALIDATION","PEXT","Payment Extension"),
 ("RS_SUMBILL_VALIDATION","SUMBILL","Summary Billing"),("RS_TPN_VALIDATION","TPN","Third-Party Notification"),
 ("RS_LIHEAP_VALIDATION","LIHEAP","LIHEAP"),("RS_PIPP_VALIDATION","PIPP","PIPP"),
 ("RS_AMP_VALIDATION","AMP","Arrearage Management"),("RS_CARE_VALIDATION","CARE","CARE/FERA"),
 ("RS_MED_VALIDATION","MED","Medical Alert"),("RS_GREEN_VALIDATION","GREEN","Green Power"),
 ("RS_CSOL_VALIDATION","CSOL","Community Solar"),("RS_ALERTS_VALIDATION","ALERTS","Usage/High-Bill Alerts"),
 ("RS_INTR_VALIDATION","INTR","Interruptible Tariff"),("RS_ADR_VALIDATION","ADR","OpenADR"),
 ("RS_CDR_VALIDATION","CDR","C&I Demand Response"),
]
sheet("11_Rulesets", "Rulesets (28 — assigned to F_EC_ELIGIBILITY_ALL)",
 "Every ruleset: precondition PROGRAM_FILTER empty OR contains <PROGRAM_ID>; standard rule sequence RU_<PGM>_010_EXCLUSIONS → 020_HARDSTOPS (DT_PROGRAM_CRITERIA) → 030_WAITING → 040_WARNINGS (DT_PROGRAM_WARNINGS) → 050_PREACTIVE (DT_PREACTIVE_INFO) → 900_LOG. Rules without content for a program are omitted.",
 ["RULESET","PROGRAM_ID","PROGRAM"],
 [[a, b, c] for a, b, c in programs],
 [26, 12, 40])

# ---------- 12 Messages ----------
msgs = [
 ("002","E","Program &1 is restricted to residential rate class"),
 ("003","E","Program &1 is restricted to commercial rate class"),
 ("012","E","Active installment plan already exists on contract account &1"),
 ("018","E","AMI meter communication disabled — open meter ticket &1 must be resolved first"),
 ("031","W","&1 returned payments in last 12 months — supervisor approval required"),
 ("044","E","Program &1 requires an AMI meter — premise has non-AMI device"),
 ("045","E","Re-enrollment waiting period active until &1 (previous participation ended &2)"),
 ("046","E","Waiting period after plan default active until &1"),
 ("052","E","Prepayment not permitted — active Medical Alert / life-support equipment at premise"),
 ("055","E","Not compatible with active Prepay service"),
 ("056","E","Prepay and Budget Billing are mutually exclusive"),
 ("057","E","Prepay requires settlement of the active payment plan first"),
 ("058","E","Only one dynamic rate program allowed — &1 is active"),
 ("062","I","Not yet eligible — requires 12 months billing history at premise (&1 of 12)"),
 ("071","E","No eligible arrears — program requires a qualifying past-due balance"),
 ("072","E","No qualifying arrears for a payment plan"),
 ("073","E","A payment extension is already active until &1"),
 ("074","E","Summary billing requires at least two contract accounts"),
 ("080","E","Curtailable load &1 kW below program minimum &2 kW"),
 ("090","I","Not yet eligible — requires an active contract with billing history"),
 ("093","I","Device installation is scheduled after contract activation"),
 ("320","W","Vehicle/charger telematics compatibility not yet verified"),
 ("321","W","Rebate subject to invoice verification"),
 ("322","W","No valid e-mail address on business partner"),
 ("330","W","&1 defaulted payment plans in last 24 months"),
 ("331","W","&1 payment extensions already granted in last 12 months"),
 ("332","W","Income verification pending — enrollment provisional"),
 ("333","W","Physician certification required within 30 days"),
 ("334","W","Community solar array at capacity — waitlist position assigned"),
 ("335","W","Interval telemetry certification pending"),
 ("340","W","Enrollment replaces the active dynamic rate program &1"),
 ("341","W","Review stacking with active subscription &1"),
 ("342","W","No Wi-Fi at premise — installation survey required"),
]
sheet("12_Messages", "Message Class ZEC_ELIG (33 messages)",
 "Severity is applied by the rules (EX_SEVERITY_MODE may downgrade E to I in pre-active mode); & = placeholder.",
 ["MSGNO","SEV","TEXT (EN)"],
 [[a, b, c] for a, b, c in msgs],
 [8, 6, 80])

out = r"C:\Users\jnamm\OneDrive\Desktop\Service Cloud for Utilities\Enrollment Center\Enrollment Center - BRFplus Configuration Workbook v1.0.xlsx"
wb.save(out)
print("Saved:", out, "| sheets:", wb.sheetnames)
